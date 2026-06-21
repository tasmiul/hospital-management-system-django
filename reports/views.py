import csv
from datetime import timedelta
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from django.http import HttpResponse
from django.db.models import Sum, Count, Q, F
from django.db.models.functions import TruncDate, TruncMonth
from django.utils import timezone
from .forms import ReportFilterForm


def export_to_csv(response, headers, rows):
    writer = csv.writer(response)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
    return response


def _date_range(request):
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    return date_from, date_to


def _apply_date_filter(queryset, date_from, date_to, field='created_at__date'):
    if date_from:
        queryset = queryset.filter(**{f'{field}__gte': date_from})
    if date_to:
        queryset = queryset.filter(**{f'{field}__lte': date_to})
    return queryset


def _date_labels(days=14):
    today = timezone.now().date()
    return [today - timedelta(days=i) for i in range(days - 1, -1, -1)]


@login_required
def report_dashboard_view(request):
    today = timezone.now().date()
    month_start = today.replace(day=1)

    from patients.models import Patient
    from appointments.models import Appointment
    from billing.models import Invoice
    from pharmacy.models import Medicine
    from laboratory.models import LabOrder
    from employees.models import Employee

    context = {
        'total_patients': Patient.objects.count(),
        'new_patients_month': Patient.objects.filter(created_at__date__gte=month_start).count(),
        'total_appointments': Appointment.objects.count(),
        'today_appointments': Appointment.objects.filter(appointment_date=today).count(),
        'total_revenue': Invoice.objects.filter(status='Paid').aggregate(total=Sum('net_amount'))['total'] or 0,
        'paid_amount': Invoice.objects.filter(status='Paid').aggregate(total=Sum('paid_amount'))['total'] or 0,
        'pending_dues': Invoice.objects.filter(status__in=['Pending', 'Partial']).aggregate(total=Sum('due_amount'))['total'] or 0,
        'total_medicines': Medicine.objects.count(),
        'low_stock_count': Medicine.objects.filter(stock_quantity__lte=F('minimum_stock')).count(),
        'total_lab_orders': LabOrder.objects.count(),
        'pending_lab_orders': LabOrder.objects.filter(status='Pending').count(),
        'total_employees': Employee.objects.filter(is_active=True).count(),
        'chart_labels': ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun'],
        'chart_data': [0, 0, 0, 0, 0, 0],
    }
    return render(request, 'reports/dashboard.html', context)


@login_required
def revenue_report_view(request):
    date_from, date_to = _date_range(request)
    from billing.models import Invoice, Payment

    invoices = Invoice.objects.all()
    invoices = _apply_date_filter(invoices, date_from, date_to)

    total_revenue = invoices.aggregate(total=Sum('net_amount'))['total'] or 0
    paid_amount = invoices.filter(status='Paid').aggregate(total=Sum('paid_amount'))['total'] or 0
    due_amount = invoices.filter(status__in=['Pending', 'Partial']).aggregate(total=Sum('due_amount'))['total'] or 0
    invoice_count = invoices.count()

    by_department = []
    from departments.models import Department
    for dept in Department.objects.all():
        dept_total = invoices.filter(
            appointment__department=dept
        ).aggregate(total=Sum('net_amount'))['total'] or 0
        if dept_total > 0:
            by_department.append({'name': dept.name, 'total': dept_total})

    dates = _date_labels()
    revenue_by_day = invoices.annotate(day=TruncDate('created_at')).values('day').annotate(total=Sum('net_amount'))
    revenue_map = {row['day']: float(row['total'] or 0) for row in revenue_by_day}

    form = ReportFilterForm(request.GET)
    context = {
        'form': form,
        'date_from': date_from,
        'date_to': date_to,
        'total_revenue': total_revenue,
        'collected': paid_amount,
        'pending': due_amount,
        'invoice_count': invoice_count,
        'paid_amount': paid_amount,
        'due_amount': due_amount,
        'by_department': by_department,
        'chart_labels': [d.strftime('%d %b') for d in dates],
        'chart_values': [revenue_map.get(d, 0) for d in dates],
        'dept_labels': [row['name'] for row in by_department],
        'dept_values': [float(row['total']) for row in by_department],
        'invoices': invoices[:200],
        'table_id': 'revenueTable',
        'report_type': 'revenue',
    }
    return render(request, 'reports/revenue.html', context)


@login_required
def appointment_report_view(request):
    date_from, date_to = _date_range(request)
    from appointments.models import Appointment

    queryset = Appointment.objects.select_related('patient', 'doctor', 'department').all()
    queryset = _apply_date_filter(queryset, date_from, date_to, field='appointment_date')

    status_filter = request.GET.get('status', '')
    if status_filter:
        queryset = queryset.filter(status=status_filter)

    total = queryset.count()
    completed = queryset.filter(status='Completed').count()
    pending = queryset.filter(status__in=['Scheduled', 'Confirmed', 'In-Progress']).count()
    cancelled = queryset.filter(status='Cancelled').count()
    no_show = queryset.filter(status='No-Show').count()

    dates = _date_labels()
    appointments_by_day = queryset.values('appointment_date').annotate(count=Count('id'))
    appointment_map = {row['appointment_date']: row['count'] for row in appointments_by_day}
    by_department = queryset.values('department__name').annotate(count=Count('id')).order_by('-count')

    form = ReportFilterForm(request.GET)
    context = {
        'form': form,
        'date_from': date_from,
        'date_to': date_to,
        'total': total,
        'total_appointments': total,
        'completed': completed,
        'pending': pending,
        'cancelled': cancelled,
        'no_show': no_show,
        'chart_labels': [d.strftime('%d %b') for d in dates],
        'chart_values': [appointment_map.get(d, 0) for d in dates],
        'dept_labels': [row['department__name'] or 'Unassigned' for row in by_department],
        'dept_values': [row['count'] for row in by_department],
        'appointments': queryset[:200],
        'table_id': 'appointmentTable',
        'report_type': 'appointment',
        'status_choices': Appointment.STATUS_CHOICES,
        'selected_status': status_filter,
    }
    return render(request, 'reports/appointment.html', context)


@login_required
def patient_report_view(request):
    date_from, date_to = _date_range(request)
    from patients.models import Patient

    queryset = Patient.objects.select_related('user').all()
    total_patients = queryset.count()

    new_patients = queryset
    if date_from:
        new_patients = new_patients.filter(created_at__date__gte=date_from)
    if date_to:
        new_patients = new_patients.filter(created_at__date__lte=date_to)
    new_count = new_patients.count()
    male_count = queryset.filter(user__gender='Male').count()
    female_count = queryset.filter(user__gender='Female').count()

    by_department = []
    from departments.models import Department
    for dept in Department.objects.all():
        dept_count = Patient.objects.filter(
            appointments__department=dept
        ).distinct().count()
        if dept_count > 0:
            by_department.append({'name': dept.name, 'count': dept_count})

    month_rows = queryset.annotate(month=TruncMonth('created_at')).values('month').annotate(count=Count('id')).order_by('month')
    age_buckets = {'0-18': 0, '19-35': 0, '36-50': 0, '51-65': 0, '65+': 0}
    today = timezone.now().date()
    for patient in queryset:
        if not patient.user.date_of_birth:
            continue
        age = today.year - patient.user.date_of_birth.year - ((today.month, today.day) < (patient.user.date_of_birth.month, patient.user.date_of_birth.day))
        if age <= 18:
            age_buckets['0-18'] += 1
        elif age <= 35:
            age_buckets['19-35'] += 1
        elif age <= 50:
            age_buckets['36-50'] += 1
        elif age <= 65:
            age_buckets['51-65'] += 1
        else:
            age_buckets['65+'] += 1

    form = ReportFilterForm(request.GET)
    context = {
        'form': form,
        'date_from': date_from,
        'date_to': date_to,
        'total_patients': total_patients,
        'new_patients': new_count,
        'new_patients_count': new_count,
        'male_count': male_count,
        'female_count': female_count,
        'by_department': by_department,
        'chart_labels': [row['month'].strftime('%b %Y') for row in month_rows],
        'chart_values': [row['count'] for row in month_rows],
        'age_labels': list(age_buckets.keys()),
        'age_values': list(age_buckets.values()),
        'patients': new_patients[:200],
        'table_id': 'patientTable',
        'report_type': 'patient',
    }
    return render(request, 'reports/patient.html', context)


@login_required
def pharmacy_report_view(request):
    date_from, date_to = _date_range(request)
    from pharmacy.models import Medicine, PrescriptionItem, DispensedMedicine

    medicines = Medicine.objects.all()
    total_medicines = medicines.count()
    total_stock_value = medicines.aggregate(
        total=Sum(F('stock_quantity') * F('unit_price'))
    )['total'] or 0

    dispensed = DispensedMedicine.objects.all()
    dispensed = _apply_date_filter(dispensed, date_from, date_to, field='dispensed_at__date')
    total_sales = dispensed.aggregate(total=Sum('total_cost'))['total'] or 0

    total_prescriptions = dispensed.count()
    low_stock_count = medicines.filter(stock_quantity__lte=F('minimum_stock')).count()
    expired_count = medicines.filter(expiry_date__lt=timezone.now().date()).count()
    dates = _date_labels()
    sales_by_day = dispensed.annotate(day=TruncDate('dispensed_at')).values('day').annotate(total=Sum('total_cost'))
    sales_map = {row['day']: float(row['total'] or 0) for row in sales_by_day}
    top_rows = PrescriptionItem.objects.values('medicine__name').annotate(qty=Sum('quantity')).order_by('-qty')[:10]
    top_medicines = medicines.order_by('-stock_quantity')[:10]
    low_stock = medicines.filter(stock_quantity__lte=F('minimum_stock'))

    form = ReportFilterForm(request.GET)
    context = {
        'form': form,
        'date_from': date_from,
        'date_to': date_to,
        'total_medicines': total_medicines,
        'total_sales': total_sales,
        'total_prescriptions': total_prescriptions,
        'low_stock_count': low_stock_count,
        'expired_count': expired_count,
        'total_stock_value': total_stock_value,
        'top_medicines': top_medicines,
        'chart_labels': [d.strftime('%d %b') for d in dates],
        'chart_values': [sales_map.get(d, 0) for d in dates],
        'top_labels': [row['medicine__name'] or 'Unknown' for row in top_rows],
        'top_values': [row['qty'] or 0 for row in top_rows],
        'low_stock': low_stock,
        'table_id': 'pharmacyTable',
        'report_type': 'pharmacy',
    }
    return render(request, 'reports/pharmacy.html', context)


@login_required
def inventory_report_view(request):
    date_from, date_to = _date_range(request)
    from inventory.models import InventoryItem, PurchaseOrder

    items = InventoryItem.objects.all()
    total_items = items.count()
    total_value = items.aggregate(
        total=Sum(F('quantity') * F('unit_price'))
    )['total'] or 0
    low_stock_items = items.filter(quantity__lte=F('reorder_level'))
    low_stock_count = low_stock_items.count()
    out_of_stock = items.filter(quantity=0).count()
    category_rows = items.values('category__name').annotate(total=Sum('quantity')).order_by('category__name')

    purchases = PurchaseOrder.objects.all()
    purchases = _apply_date_filter(purchases, date_from, date_to, field='order_date')
    total_purchases = purchases.aggregate(total=Sum('total_cost'))['total'] or 0
    purchase_count = purchases.count()

    form = ReportFilterForm(request.GET)
    context = {
        'form': form,
        'date_from': date_from,
        'date_to': date_to,
        'total_items': total_items,
        'total_value': total_value,
        'low_stock_items': low_stock_items,
        'low_stock_count': low_stock_count,
        'out_of_stock': out_of_stock,
        'cat_labels': [row['category__name'] or 'Uncategorized' for row in category_rows],
        'cat_values': [row['total'] or 0 for row in category_rows],
        'total_purchases': total_purchases,
        'purchase_count': purchase_count,
        'purchases': purchases[:200],
        'table_id': 'inventoryTable',
        'report_type': 'inventory',
    }
    return render(request, 'reports/inventory.html', context)


@login_required
def lab_report_view(request):
    date_from, date_to = _date_range(request)
    from laboratory.models import LabOrder, LabTest

    orders = LabOrder.objects.select_related('patient', 'doctor').all()
    orders = _apply_date_filter(orders, date_from, date_to)

    total_tests = orders.aggregate(total=Count('items'))['total'] or 0
    total_orders = orders.count()
    completed = orders.filter(status='Completed').count()
    pending = orders.filter(status='Pending').count()
    in_progress = orders.filter(status='In-Progress').count()

    tests = LabTest.objects.all()
    total_revenue = orders.filter(status='Completed').aggregate(total=Sum('items__test__price'))['total'] or 0
    dates = _date_labels()
    orders_by_day = orders.annotate(day=TruncDate('created_at')).values('day').annotate(count=Count('id'))
    order_map = {row['day']: row['count'] for row in orders_by_day}
    top_tests = LabTest.objects.filter(laborderitem__order__in=orders).values('name').annotate(count=Count('laborderitem')).order_by('-count')[:10]

    form = ReportFilterForm(request.GET)
    context = {
        'form': form,
        'date_from': date_from,
        'date_to': date_to,
        'total_orders': total_orders,
        'total_tests': total_tests,
        'completed': completed,
        'pending': pending,
        'in_progress': in_progress,
        'total_revenue': total_revenue,
        'chart_labels': [d.strftime('%d %b') for d in dates],
        'chart_values': [order_map.get(d, 0) for d in dates],
        'top_labels': [row['name'] for row in top_tests],
        'top_values': [row['count'] for row in top_tests],
        'orders': orders[:200],
        'table_id': 'labTable',
        'report_type': 'lab',
    }
    return render(request, 'reports/lab.html', context)


@login_required
def employee_report_view(request):
    from employees.models import Employee, Attendance
    from departments.models import Department

    department_filter = request.GET.get('department', '')
    employees = Employee.objects.select_related('user', 'department').filter(is_active=True)

    if department_filter:
        employees = employees.filter(department__name=department_filter)

    total_employees = employees.count()
    active_employees = total_employees
    doctor_count = employees.filter(user__roles__name='Doctor').distinct().count()
    nurse_count = employees.filter(user__roles__name='Nurse').distinct().count()

    departments = Department.objects.all()
    dept_summary = []
    for dept in departments:
        count = employees.filter(department=dept).count()
        if count > 0:
            dept_summary.append({'name': dept.name, 'count': count})

    today = timezone.now().date()
    month_start = today.replace(day=1)
    attendance_stats = Attendance.objects.filter(date__gte=month_start).values('status').annotate(
        count=Count('id')
    )
    desig_rows = employees.values('designation').annotate(count=Count('id')).order_by('-count')

    form = ReportFilterForm(request.GET)
    context = {
        'form': form,
        'total_employees': total_employees,
        'active_employees': active_employees,
        'doctor_count': doctor_count,
        'nurse_count': nurse_count,
        'dept_summary': dept_summary,
        'dept_labels': [row['name'] for row in dept_summary],
        'dept_values': [row['count'] for row in dept_summary],
        'desig_labels': [row['designation'] for row in desig_rows],
        'desig_values': [row['count'] for row in desig_rows],
        'attendance_stats': list(attendance_stats),
        'employees': employees[:200],
        'table_id': 'employeeTable',
        'report_type': 'employee',
    }
    return render(request, 'reports/employee.html', context)


@login_required
def export_revenue_pdf_view(request):
    from billing.models import Invoice
    from django.http import HttpResponse
    from xhtml2pdf import pisa
    from io import BytesIO

    date_from, date_to = _date_range(request)
    invoices = Invoice.objects.all()
    invoices = _apply_date_filter(invoices, date_from, date_to)

    total_revenue = invoices.aggregate(total=Sum('net_amount'))['total'] or 0
    paid_amount = invoices.filter(status='Paid').aggregate(total=Sum('paid_amount'))['total'] or 0
    due_amount = invoices.filter(status__in=['Pending', 'Partial']).aggregate(total=Sum('due_amount'))['total'] or 0

    html_content = f"""
    <html>
    <head><style>
        body {{ font-family: Helvetica, sans-serif; font-size: 12px; }}
        h1 {{ color: #333; text-align: center; }}
        table {{ width: 100%; border-collapse: collapse; margin-top: 20px; }}
        th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
        th {{ background-color: #4CAF50; color: white; }}
        .summary {{ margin: 20px 0; padding: 10px; background: #f9f9f9; border-radius: 5px; }}
    </style></head>
    <body>
        <h1>Revenue Report</h1>
        <div class="summary">
            <p><strong>Total Revenue:</strong> ${total_revenue:,.2f}</p>
            <p><strong>Paid Amount:</strong> ${paid_amount:,.2f}</p>
            <p><strong>Due Amount:</strong> ${due_amount:,.2f}</p>
        </div>
        <table>
            <tr><th>Invoice #</th><th>Patient</th><th>Net Amount</th><th>Paid</th><th>Due</th><th>Status</th></tr>
    """
    for inv in invoices[:100]:
        html_content += f"""
            <tr>
                <td>{inv.invoice_number}</td>
                <td>{inv.patient}</td>
                <td>${inv.net_amount:,.2f}</td>
                <td>${inv.paid_amount:,.2f}</td>
                <td>${inv.due_amount:,.2f}</td>
                <td>{inv.status}</td>
            </tr>
        """
    html_content += "</table></body></html>"

    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html_content.encode('utf-8')), result)
    if not pdf.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="revenue_report.pdf"'
        return response
    return HttpResponse('Error generating PDF', status=500)


@login_required
def export_revenue_excel_view(request):
    from billing.models import Invoice
    from openpyxl import Workbook
    from io import BytesIO

    date_from, date_to = _date_range(request)
    invoices = Invoice.objects.all()
    invoices = _apply_date_filter(invoices, date_from, date_to)

    wb = Workbook()
    ws = wb.active
    ws.title = "Revenue Report"

    ws.append(['Invoice #', 'Patient', 'Total Amount', 'Discount', 'Tax', 'Net Amount', 'Paid', 'Due', 'Status', 'Date'])

    for inv in invoices:
        ws.append([
            inv.invoice_number,
            str(inv.patient),
            float(inv.total_amount),
            float(inv.discount),
            float(inv.tax),
            float(inv.net_amount),
            float(inv.paid_amount),
            float(inv.due_amount),
            inv.status,
            inv.created_at.strftime('%Y-%m-%d'),
        ])

    total_row = ws.max_row + 2
    ws.cell(row=total_row, column=1, value='Summary')
    ws.cell(row=total_row + 1, column=1, value='Total Revenue')
    ws.cell(row=total_row + 1, column=2, value=float(invoices.aggregate(t=Sum('net_amount'))['t'] or 0))
    ws.cell(row=total_row + 2, column=1, value='Paid')
    ws.cell(row=total_row + 2, column=2, value=float(invoices.filter(status='Paid').aggregate(t=Sum('paid_amount'))['t'] or 0))
    ws.cell(row=total_row + 3, column=1, value='Due')
    ws.cell(row=total_row + 3, column=2, value=float(invoices.filter(status__in=['Pending', 'Partial']).aggregate(t=Sum('due_amount'))['t'] or 0))

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="revenue_report.xlsx"'
    return response


@login_required
def export_appointment_excel_view(request):
    from appointments.models import Appointment
    from openpyxl import Workbook
    from io import BytesIO

    date_from, date_to = _date_range(request)
    queryset = Appointment.objects.select_related('patient', 'doctor', 'department').all()
    queryset = _apply_date_filter(queryset, date_from, date_to, field='appointment_date')

    wb = Workbook()
    ws = wb.active
    ws.title = "Appointment Report"

    ws.append(['Patient', 'Doctor', 'Department', 'Date', 'Time', 'Type', 'Status'])

    for apt in queryset:
        ws.append([
            str(apt.patient),
            str(apt.doctor),
            str(apt.department) if apt.department else '',
            apt.appointment_date.strftime('%Y-%m-%d'),
            apt.appointment_time.strftime('%H:%M'),
            apt.appointment_type,
            apt.status,
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="appointment_report.xlsx"'
    return response


@login_required
def export_patient_excel_view(request):
    from patients.models import Patient
    from openpyxl import Workbook
    from io import BytesIO

    date_from, date_to = _date_range(request)
    queryset = Patient.objects.select_related('user').all()
    queryset = _apply_date_filter(queryset, date_from, date_to)

    wb = Workbook()
    ws = wb.active
    ws.title = "Patient Report"

    ws.append(['Patient ID', 'Name', 'Email', 'Phone', 'Blood Group', 'Date Joined'])

    for patient in queryset:
        ws.append([
            patient.patient_id,
            patient.user.get_full_name(),
            patient.user.email,
            patient.user.phone,
            patient.blood_group,
            patient.created_at.strftime('%Y-%m-%d'),
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="patient_report.xlsx"'
    return response


@login_required
def export_revenue_csv_view(request):
    from billing.models import Invoice
    date_from, date_to = _date_range(request)
    invoices = Invoice.objects.all()
    invoices = _apply_date_filter(invoices, date_from, date_to)

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="revenue_report.csv"'
    headers = ['Invoice #', 'Patient', 'Net Amount', 'Paid', 'Due', 'Status', 'Date']
    rows = []
    for inv in invoices:
        rows.append([
            inv.invoice_number, str(inv.patient), float(inv.net_amount),
            float(inv.paid_amount), float(inv.due_amount), inv.status,
            inv.created_at.strftime('%Y-%m-%d'),
        ])
    return export_to_csv(response, headers, rows)


@login_required
def export_appointment_csv_view(request):
    from appointments.models import Appointment
    date_from, date_to = _date_range(request)
    queryset = Appointment.objects.select_related('patient', 'doctor', 'department').all()
    queryset = _apply_date_filter(queryset, date_from, date_to, field='appointment_date')

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="appointment_report.csv"'
    headers = ['Patient', 'Doctor', 'Department', 'Date', 'Time', 'Type', 'Status']
    rows = []
    for apt in queryset:
        rows.append([
            str(apt.patient), str(apt.doctor),
            str(apt.department) if apt.department else '',
            apt.appointment_date.strftime('%Y-%m-%d'),
            apt.appointment_time.strftime('%H:%M'),
            apt.appointment_type, apt.status,
        ])
    return export_to_csv(response, headers, rows)


@login_required
def export_patient_csv_view(request):
    from patients.models import Patient
    date_from, date_to = _date_range(request)
    queryset = Patient.objects.select_related('user').all()
    queryset = _apply_date_filter(queryset, date_from, date_to)

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="patient_report.csv"'
    headers = ['Patient ID', 'Name', 'Email', 'Phone', 'Blood Group', 'Date Joined']
    rows = []
    for patient in queryset:
        rows.append([
            patient.patient_id, patient.user.get_full_name(),
            patient.user.email, patient.user.phone,
            patient.blood_group, patient.created_at.strftime('%Y-%m-%d'),
        ])
    return export_to_csv(response, headers, rows)


@login_required
def export_pharmacy_csv_view(request):
    from pharmacy.models import Medicine, DispensedMedicine
    date_from, date_to = _date_range(request)
    medicines = Medicine.objects.all()
    dispensed = DispensedMedicine.objects.all()
    dispensed = _apply_date_filter(dispensed, date_from, date_to, field='dispensed_at__date')

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="pharmacy_report.csv"'
    headers = ['Medicine', 'Generic', 'Stock', 'Unit Price', 'Total Value', 'Status']
    rows = []
    for med in medicines:
        status = 'Low Stock' if med.stock_quantity <= med.minimum_stock else 'OK'
        rows.append([
            med.name, med.generic_name or '', med.stock_quantity,
            float(med.unit_price), float(med.stock_quantity * med.unit_price), status,
        ])
    return export_to_csv(response, headers, rows)


@login_required
def export_lab_csv_view(request):
    from laboratory.models import LabOrder
    date_from, date_to = _date_range(request)
    orders = LabOrder.objects.select_related('patient', 'doctor').all()
    orders = _apply_date_filter(orders, date_from, date_to)

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="lab_report.csv"'
    headers = ['Order #', 'Patient', 'Doctor', 'Status', 'Date']
    rows = []
    for order in orders:
        rows.append([
            order.order_number, str(order.patient), str(order.doctor),
            order.status, order.created_at.strftime('%Y-%m-%d'),
        ])
    return export_to_csv(response, headers, rows)


@login_required
def export_inventory_csv_view(request):
    from inventory.models import InventoryItem, PurchaseOrder
    date_from, date_to = _date_range(request)
    items = InventoryItem.objects.all()
    purchases = PurchaseOrder.objects.all()
    purchases = _apply_date_filter(purchases, date_from, date_to, field='order_date')

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="inventory_report.csv"'
    headers = ['Item', 'Category', 'Quantity', 'Unit Price', 'Total Value', 'Reorder Level', 'Status']
    rows = []
    for item in items:
        status = 'Low Stock' if item.quantity <= item.reorder_level else 'OK'
        rows.append([
            item.name, item.category.name if item.category else '',
            item.quantity, float(item.unit_price),
            float(item.quantity * item.unit_price),
            item.reorder_level, status,
        ])
    return export_to_csv(response, headers, rows)


@login_required
def export_employee_csv_view(request):
    from employees.models import Employee
    employees = Employee.objects.select_related('user', 'department').filter(is_active=True)

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="employee_report.csv"'
    headers = ['Employee ID', 'Name', 'Department', 'Designation', 'Phone', 'Email']
    rows = []
    for emp in employees:
        rows.append([
            emp.employee_id, emp.user.get_full_name(),
            emp.department.name if emp.department else '',
            emp.designation, emp.user.phone, emp.user.email,
        ])
    return export_to_csv(response, headers, rows)


@login_required
def export_pharmacy_excel_view(request):
    from pharmacy.models import Medicine
    from openpyxl import Workbook
    from io import BytesIO

    date_from, date_to = _date_range(request)
    medicines = Medicine.objects.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Pharmacy Report"
    ws.append(['Medicine', 'Generic', 'Category', 'Stock', 'Unit Price', 'Total Value', 'Status'])

    for med in medicines:
        status = 'Low Stock' if med.stock_quantity <= med.minimum_stock else 'OK'
        ws.append([
            med.name, med.generic_name or '',
            med.category.name if med.category else '',
            med.stock_quantity, float(med.unit_price),
            float(med.stock_quantity * med.unit_price), status,
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="pharmacy_report.xlsx"'
    return response


@login_required
def export_lab_excel_view(request):
    from laboratory.models import LabOrder
    from openpyxl import Workbook
    from io import BytesIO

    date_from, date_to = _date_range(request)
    orders = LabOrder.objects.select_related('patient', 'doctor').all()
    orders = _apply_date_filter(orders, date_from, date_to)

    wb = Workbook()
    ws = wb.active
    ws.title = "Lab Report"
    ws.append(['Order #', 'Patient', 'Doctor', 'Status', 'Date'])

    for order in orders:
        ws.append([
            order.order_number, str(order.patient), str(order.doctor),
            order.status, order.created_at.strftime('%Y-%m-%d'),
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="lab_report.xlsx"'
    return response


@login_required
def export_inventory_excel_view(request):
    from inventory.models import InventoryItem
    from openpyxl import Workbook
    from io import BytesIO

    items = InventoryItem.objects.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory Report"
    ws.append(['Item', 'Category', 'Quantity', 'Unit Price', 'Total Value', 'Reorder Level', 'Status'])

    for item in items:
        status = 'Low Stock' if item.quantity <= item.reorder_level else 'OK'
        ws.append([
            item.name, item.category.name if item.category else '',
            item.quantity, float(item.unit_price),
            float(item.quantity * item.unit_price),
            item.reorder_level, status,
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="inventory_report.xlsx"'
    return response


@login_required
def export_employee_excel_view(request):
    from employees.models import Employee
    from openpyxl import Workbook
    from io import BytesIO

    employees = Employee.objects.select_related('user', 'department').filter(is_active=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Employee Report"
    ws.append(['Employee ID', 'Name', 'Department', 'Designation', 'Phone', 'Email'])

    for emp in employees:
        ws.append([
            emp.employee_id, emp.user.get_full_name(),
            emp.department.name if emp.department else '',
            emp.designation, emp.user.phone, emp.user.email,
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="employee_report.xlsx"'
    return response


@login_required
def export_appointment_pdf_view(request):
    from appointments.models import Appointment
    from xhtml2pdf import pisa
    from io import BytesIO

    date_from, date_to = _date_range(request)
    queryset = Appointment.objects.select_related('patient', 'doctor', 'department').all()
    queryset = _apply_date_filter(queryset, date_from, date_to, field='appointment_date')

    html_content = """
    <html><head><style>
        body { font-family: Helvetica, sans-serif; font-size: 12px; }
        h1 { color: #333; text-align: center; }
        table { width: 100%; border-collapse: collapse; margin-top: 20px; }
        th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }
        th { background-color: #4CAF50; color: white; }
    </style></head><body>
        <h1>Appointment Report</h1>
        <table>
            <tr><th>Patient</th><th>Doctor</th><th>Department</th><th>Date</th><th>Time</th><th>Type</th><th>Status</th></tr>
    """
    for apt in queryset[:100]:
        html_content += f"""
            <tr>
                <td>{apt.patient}</td><td>{apt.doctor}</td>
                <td>{apt.department or ''}</td>
                <td>{apt.appointment_date.strftime('%Y-%m-%d')}</td>
                <td>{apt.appointment_time.strftime('%H:%M')}</td>
                <td>{apt.appointment_type}</td><td>{apt.status}</td>
            </tr>"""
    html_content += "</table></body></html>"

    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html_content.encode('utf-8')), result)
    if not pdf.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="appointment_report.pdf"'
        return response
    return HttpResponse('Error generating PDF', status=500)


@login_required
def export_patient_pdf_view(request):
    from patients.models import Patient
    from xhtml2pdf import pisa
    from io import BytesIO

    date_from, date_to = _date_range(request)
    queryset = Patient.objects.select_related('user').all()
    queryset = _apply_date_filter(queryset, date_from, date_to)

    html_content = """
    <html><head><style>
        body { font-family: Helvetica, sans-serif; font-size: 12px; }
        h1 { color: #333; text-align: center; }
        table { width: 100%; border-collapse: collapse; margin-top: 20px; }
        th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }
        th { background-color: #4CAF50; color: white; }
    </style></head><body>
        <h1>Patient Report</h1>
        <table>
            <tr><th>Patient ID</th><th>Name</th><th>Email</th><th>Phone</th><th>Blood Group</th><th>Date Joined</th></tr>
    """
    for patient in queryset[:100]:
        html_content += f"""
            <tr>
                <td>{patient.patient_id}</td><td>{patient.user.get_full_name()}</td>
                <td>{patient.user.email}</td><td>{patient.user.phone}</td>
                <td>{patient.blood_group}</td>
                <td>{patient.created_at.strftime('%Y-%m-%d')}</td>
            </tr>"""
    html_content += "</table></body></html>"

    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html_content.encode('utf-8')), result)
    if not pdf.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="patient_report.pdf"'
        return response
    return HttpResponse('Error generating PDF', status=500)


@login_required
def export_pharmacy_pdf_view(request):
    from pharmacy.models import Medicine
    from xhtml2pdf import pisa
    from io import BytesIO

    date_from, date_to = _date_range(request)
    medicines = Medicine.objects.all()

    html_content = """
    <html><head><style>
        body { font-family: Helvetica, sans-serif; font-size: 12px; }
        h1 { color: #333; text-align: center; }
        table { width: 100%; border-collapse: collapse; margin-top: 20px; }
        th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }
        th { background-color: #4CAF50; color: white; }
    </style></head><body>
        <h1>Pharmacy Report</h1>
        <table>
            <tr><th>Medicine</th><th>Generic</th><th>Stock</th><th>Unit Price</th><th>Total Value</th><th>Status</th></tr>
    """
    for med in medicines[:100]:
        status = 'Low Stock' if med.stock_quantity <= med.minimum_stock else 'OK'
        html_content += f"""
            <tr>
                <td>{med.name}</td><td>{med.generic_name or ''}</td>
                <td>{med.stock_quantity}</td><td>{med.unit_price}</td>
                <td>{med.stock_quantity * med.unit_price}</td><td>{status}</td>
            </tr>"""
    html_content += "</table></body></html>"

    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html_content.encode('utf-8')), result)
    if not pdf.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="pharmacy_report.pdf"'
        return response
    return HttpResponse('Error generating PDF', status=500)


@login_required
def export_lab_pdf_view(request):
    from laboratory.models import LabOrder
    from xhtml2pdf import pisa
    from io import BytesIO

    date_from, date_to = _date_range(request)
    orders = LabOrder.objects.select_related('patient', 'doctor').all()
    orders = _apply_date_filter(orders, date_from, date_to)

    html_content = """
    <html><head><style>
        body { font-family: Helvetica, sans-serif; font-size: 12px; }
        h1 { color: #333; text-align: center; }
        table { width: 100%; border-collapse: collapse; margin-top: 20px; }
        th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }
        th { background-color: #4CAF50; color: white; }
    </style></head><body>
        <h1>Lab Report</h1>
        <table>
            <tr><th>Order #</th><th>Patient</th><th>Doctor</th><th>Status</th><th>Date</th></tr>
    """
    for order in orders[:100]:
        html_content += f"""
            <tr>
                <td>{order.order_number}</td><td>{order.patient}</td>
                <td>{order.doctor}</td><td>{order.status}</td>
                <td>{order.created_at.strftime('%Y-%m-%d')}</td>
            </tr>"""
    html_content += "</table></body></html>"

    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html_content.encode('utf-8')), result)
    if not pdf.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="lab_report.pdf"'
        return response
    return HttpResponse('Error generating PDF', status=500)


@login_required
def export_inventory_pdf_view(request):
    from inventory.models import InventoryItem
    from xhtml2pdf import pisa
    from io import BytesIO

    items = InventoryItem.objects.all()

    html_content = """
    <html><head><style>
        body { font-family: Helvetica, sans-serif; font-size: 12px; }
        h1 { color: #333; text-align: center; }
        table { width: 100%; border-collapse: collapse; margin-top: 20px; }
        th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }
        th { background-color: #4CAF50; color: white; }
    </style></head><body>
        <h1>Inventory Report</h1>
        <table>
            <tr><th>Item</th><th>Category</th><th>Quantity</th><th>Unit Price</th><th>Total Value</th><th>Status</th></tr>
    """
    for item in items[:100]:
        status = 'Low Stock' if item.quantity <= item.reorder_level else 'OK'
        html_content += f"""
            <tr>
                <td>{item.name}</td><td>{item.category.name if item.category else ''}</td>
                <td>{item.quantity}</td><td>{item.unit_price}</td>
                <td>{item.quantity * item.unit_price}</td><td>{status}</td>
            </tr>"""
    html_content += "</table></body></html>"

    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html_content.encode('utf-8')), result)
    if not pdf.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="inventory_report.pdf"'
        return response
    return HttpResponse('Error generating PDF', status=500)


@login_required
def export_employee_pdf_view(request):
    from employees.models import Employee
    from xhtml2pdf import pisa
    from io import BytesIO

    employees = Employee.objects.select_related('user', 'department').filter(is_active=True)

    html_content = """
    <html><head><style>
        body { font-family: Helvetica, sans-serif; font-size: 12px; }
        h1 { color: #333; text-align: center; }
        table { width: 100%; border-collapse: collapse; margin-top: 20px; }
        th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }
        th { background-color: #4CAF50; color: white; }
    </style></head><body>
        <h1>Employee Report</h1>
        <table>
            <tr><th>Employee ID</th><th>Name</th><th>Department</th><th>Designation</th><th>Phone</th></tr>
    """
    for emp in employees[:100]:
        html_content += f"""
            <tr>
                <td>{emp.employee_id}</td><td>{emp.user.get_full_name()}</td>
                <td>{emp.department.name if emp.department else ''}</td>
                <td>{emp.designation}</td><td>{emp.user.phone}</td>
            </tr>"""
    html_content += "</table></body></html>"

    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html_content.encode('utf-8')), result)
    if not pdf.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="employee_report.pdf"'
        return response
    return HttpResponse('Error generating PDF', status=500)
